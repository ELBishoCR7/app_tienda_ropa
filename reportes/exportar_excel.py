import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime


def exportar_a_excel(nombre_archivo: str, encabezados: list, datos: list):
    """
    Exporta datos a un archivo Excel.
    
    :param nombre_archivo: Nombre del archivo de salida (ej: 'clientes.xlsx')
    :param encabezados: Lista con los nombres de las columnas
    :param datos: Lista de tuplas o listas con los registros
    """
    try:
        # Crear libro y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Escribir encabezados
        for col, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(row=1, column=col, value=encabezado)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")

        # Escribir datos
        for fila, registro in enumerate(datos, start=2):
            for col, valor in enumerate(registro, start=1):
                ws.cell(row=fila, column=col, value=valor)

        # Ajustar ancho de columnas automáticamente
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Guardar archivo con timestamp para evitar sobrescribir
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_final = f"{nombre_archivo.replace('.xlsx','')}_{timestamp}.xlsx"
        wb.save(archivo_final)

        return archivo_final

    except Exception as e:
        print(f"Error al exportar a Excel: {e}")
        return None